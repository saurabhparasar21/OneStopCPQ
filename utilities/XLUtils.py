import openpyxl

def getRowCount(file,SheetName):
    workboook=openpyxl.load_workbook(file)
    sheet=workboook[SheetName]
    return(sheet.max_row)

def getColumnCount(file,SheetName):
    workbook=openpyxl.load_workbook(file)
    sheet=workbook[SheetName]
    return(sheet.max_column)

def readData(file,SheetName,rowno,colno):
    workbook=openpyxl.load_workbook(file)
    sheet=workbook[SheetName]
    return sheet.cell(row=rowno,column=colno).value

def writeData(file,SheetName,rowno,colno,Data):
    workbook=openpyxl.load_workbook(file)
    sheet=workbook[SheetName]
    Update=sheet.cell(row=rowno,column=colno)
    Update.value=Data
    workbook.save(file)